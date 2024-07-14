import openpyxl
import os
from tkinter import filedialog as fd
def get_file():
    name = fd.askopenfilename()
    return name


file_path=get_file()
if (file_path[-5:] != '.xlsx' and file_path[-4:] != '.xls'):
    exit(1)
file = openpyxl.load_workbook(file_path)
table = file.active
max_col=table.max_column+1
max_row=table.max_row+1
asortiment=dict()
adresses=list()
codes = list()
for i in range(2,max_row):
    asortiment[table.cell(row=i,column=1).value]=table.cell(row=i,column=2).value
    codes.append (table.cell(row=i,column=1).value)

for i in range(3,max_col):
    adresses.append(table.cell(row=1,column=i).value)

matrix = list()
for i in range (3,max_col):
    col=list()
    for j in range (2,max_row):
        val=table.cell(row=j,column=i).value
        if (val==1):
            col.append(1)
        else:
            col.append(0)
    matrix.append(col)











direct = file_path.split('.')[0]

os.mkdir(direct)
k=0
for file in adresses:
    f = open(direct+"/"+file.replace('/','_')+".txt", "w")

    text=""
    t=0
    for i in codes:
        if matrix[k][t]==1:
            text=text+str(i)+"\n"
        t+=1
    f.write(text)
    f.close()
    k+=1
